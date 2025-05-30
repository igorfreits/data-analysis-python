import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill,Alignment
from openpyxl import load_workbook
import os
 
# Função para converter .xls para .xlsx
import pandas as pd
import os


def converter_xls_para_xlsx(caminho_arquivo):
    if not caminho_arquivo or not isinstance(caminho_arquivo, str):
        raise ValueError("Caminho do arquivo inválido.")

    if not caminho_arquivo.lower().endswith(".xls"):
        raise ValueError("O arquivo fornecido não é um .xls")

    if not os.path.exists(caminho_arquivo):
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho_arquivo}")

    novo_caminho = caminho_arquivo.replace(".xls", ".xlsx")

    try:
        df = pd.read_excel(caminho_arquivo, engine="xlrd")
        if df.empty:
            raise ValueError("O arquivo está vazio.")

        df.to_excel(novo_caminho, index=False, engine="openpyxl")
        return novo_caminho
    except Exception as e:
        print(f"Erro ao converter {caminho_arquivo}: {e}")
        return None



# Caminho dos arquivos
data_path = 'data-analysis-python/PROCESSADO ERRO/'

# Caminho do arquivo original
arquivo_xls = data_path + 'Benner - Processado Erro 0.xls'

# Verifica se o arquivo .xls existe e converte antes de ler
if os.path.exists(arquivo_xls):
    arquivo_xls = converter_xls_para_xlsx(arquivo_xls)

# ARQUIVOS
processado_erro = pd.read_excel(arquivo_xls)
realocacao = pd.read_excel(data_path + 'Realocacao.xlsx')
 
parametros, list_erros, info, clientes_fcm = [
    pd.read_excel(data_path + 'Parametros.xlsx', sheet_name=sheet)
   
    for sheet in ['Parametros', 'Lista de erros', 'Info', 'Clientes FCM']]
 
# Declaração de guias - Relatório DASH
wb = load_workbook(data_path + 'Relatorio - Dash.xlsx')
relatorio_base = wb['Processado Erro - BASE']
 
# Limpeza de planilha
relatorio_base.delete_rows(2, relatorio_base.max_row)
relatorio_base.delete_cols(2, relatorio_base.max_column)
 
# Criação de colunas
processado_erro['Aging Alteração'] = ''
processado_erro['Aging Inclusão'] = ''
processado_erro['OBTS'] = ''
processado_erro['CAMPO'] = 'Não identificado'
processado_erro['ORIGEM DO ERRO'] = 'Análise Benner'
processado_erro['TIPO DE ERRO'] = 'Sistema'
processado_erro['CATEGORIA DE ERRO'] = 'Qualidade dos dados'
processado_erro['EMPRESA'] = ''
processado_erro['RESPONSÁVEL'] = 'Operações - CORP'
 
# Preenchimento de valores nulos
processado_erro['Mensagem Erro'].fillna(
    'Erro não localizado', inplace=True)
 
processado_erro.fillna('-', inplace=True)
  
# Inserção dos OBTS
processado_erro['OBTS'] = processado_erro['OBT']
processado_erro['OBTS'] = processado_erro['OBTS'].str.replace(
    'TMS', 'ARGO(TMS)')
 
mask_zupper = (processado_erro['Canal de Vendas'].str.contains('ZUPPER', case=False) |
               processado_erro['Grupo Empresarial'].str.contains('Zupper', case=False))
 
processado_erro['OBTS'] = processado_erro['OBTS'].mask(mask_zupper, 'ZUPPER')
 
mask_kontrip = processado_erro['Canal de Vendas'].str.contains(
    'KONTRIP', case=False)
processado_erro['OBTS'] = processado_erro['OBTS'].mask(mask_kontrip, 'KONTRIP')
 
processado_erro['Mensagem Erro'] = processado_erro['Mensagem Erro'].astype(str)
parametros['Mensagem'] = parametros['Mensagem'].astype(str)
 
# Atribuição de campos, origem do erro e tipo de erro
for row in range(len(processado_erro)):
    for row2 in range(len(parametros)):
        if parametros['Mensagem'][row2] in processado_erro['Mensagem Erro'][row]:
            processado_erro.at[row, 'CAMPO'] = parametros.at[row2, 'Campo']
            processado_erro.at[row, 'ORIGEM DO ERRO'] = parametros.at[row2, 'Origem do Erro']
            processado_erro.at[row, 'TIPO DE ERRO'] = parametros.at[row2, 'Tipo de Erro']
            processado_erro.at[row, 'CATEGORIA DE ERRO'] = parametros.at[row2, 'Categoria de Erro']

# Formatação de Datas
processado_erro['Aging Inclusão'] = (
    datetime.now() - pd.to_datetime(processado_erro['Data Inclusão'].str[:10], format='%d/%m/%Y')).dt.days

processado_erro['Aging Alteração'] = (
    datetime.now() - pd.to_datetime(processado_erro['Data Alteração'].str[:10], format='%d/%m/%Y')).dt.days

# Leitura de data de alteração - Edição não permitida
processado_erro.loc[processado_erro['TIPO DE ERRO'].str.contains('Edição não Permitida'), 'Aging Inclusão'] = (
    datetime.now() - pd.to_datetime(processado_erro['Data Alteração'].str[:10], format='%d/%m/%Y')
).dt.days

# Leitura de data de alteração - Bilhete duplicado
processado_erro.loc[processado_erro['CAMPO'].str.contains('Bilhete duplicado'), 'Aging Inclusão'] = (
    datetime.now() - pd.to_datetime(processado_erro['Data Alteração'].str[:10], format='%d/%m/%Y')
).dt.days

processado_erro.loc[processado_erro['Data Emissão'].str.contains(
    '-'), 'Data Emissão'] = processado_erro['Data Inclusão']
 
# Categorização de dias parados
limites = [0, 3, 6, 9, 16, 24, 31, float('inf')]
rotulos = ['0 a 02 dias', '03 a 05 dias', '06 a 08 dias', '09 a 15 dias', '16 a 23 dias', '24 a 31 dias', '31 dias ou +']
 
processado_erro['Aging Inclusão'] = pd.cut(processado_erro['Aging Inclusão'], bins=limites, labels=rotulos, right=False, include_lowest=True)
processado_erro['Aging Alteração'] = pd.cut(processado_erro['Aging Alteração'], bins=limites, labels=rotulos, right=False, include_lowest=True)

#formatação de datas
# Aplicar formatação com dayfirst=True para ajustar o formato de data "DD/MM/AAAA HH:MM"
processado_erro['Data Inclusão'] = pd.to_datetime(processado_erro['Data Inclusão'], format='mixed', dayfirst=True, errors='coerce')
processado_erro['Data Emissão'] = pd.to_datetime(processado_erro['Data Emissão'], format='mixed', dayfirst=True, errors='coerce')
processado_erro['Data Alteração'] = pd.to_datetime(processado_erro['Data Alteração'], format='mixed', dayfirst=True, errors='coerce')

# Atribuição - Fornecedor
processado_erro.loc[processado_erro['Mensagem Erro'].str.contains('Fornecedor não preenchido', case=False), 
                    ['CAMPO', 'ORIGEM DO ERRO', 'TIPO DE ERRO']] = ['Falta de Fornecedor', 'Campo Fornecedor', 'Dados do Fornecedor']

# Atribuição - de responsáveis
for row in range(len(processado_erro)):
    for row2 in range(len(info)):
        if info['CAMPO_INFO'][row2] in processado_erro['CAMPO'][row]:
            processado_erro.at[row, 'RESPONSÁVEL'] = info.at[row2, 'RESPONSÁVEL_INFO']


# Preenchimento de valores nulos(Status Requisição)
condition = processado_erro['Status Requisicao'].str.contains('-', na=False)
processado_erro.loc[condition, 'Status Requisicao'] = 'OFF LINE'
 
# Preenchimento de vendas manuais
condition = processado_erro['OBT'].str.contains('MANUAL', na=False)
processado_erro.loc[condition, 'Status Requisicao'] = 'OFF LINE'

# Preenchimento de valores repetidos - falta de informação gerencial
processado_erro.loc[processado_erro['Mensagem Erro'].str.count('não preenchid')
                    > 1, 'ORIGEM DO ERRO'] = 'Mais de um campo não preenchido'

# Realocações - Suporte KCS (Falta de informação Gerencial e SABRE)
processado_erro.loc[
    (processado_erro['CAMPO'].str.contains('Falta de informação Gerencial')),
    'RESPONSÁVEL'] = 'Suporte KCS'
 
# Realocações - Operações - CORP (Reembolsos Recebidos)
processado_erro.loc[
    (processado_erro['Cliente'] == 'Reembolsos Recebidos'),
    'RESPONSÁVEL'] = 'Operações - CORP'
 
# Realocações - Operações - CORP (WS, WT, CT - MANUAL)
processado_erro.loc[
    (processado_erro['OBTS'] == 'MANUAL') &
    (processado_erro['Agente Emissão'] == 'WS') |
    (processado_erro['Agente Emissão'] == 'WT') |
    (processado_erro['Agente Emissão'] == 'CT'),
    'RESPONSÁVEL'] = 'Operações - CORP'
 
# Realocações - Operações - CORP (MANUAL e Accounting sem trecho)
processado_erro.loc[
    (processado_erro['OBTS'] == 'MANUAL') &
    (processado_erro['Mensagem Erro'].str.contains('Accounting sem trecho')),
    'RESPONSÁVEL'] = 'Operações - CORP'
 
# Realocação - FCM
clientes_fcm_str = '|'.join(map(str, clientes_fcm['Clientes FCM']))
 
processado_erro['RESPONSÁVEL'] = processado_erro['Cliente'].apply(
    lambda cliente: 'Operações - FCM' if any(cliente_fcm in cliente for cliente_fcm in clientes_fcm['Clientes FCM']) else processado_erro['RESPONSÁVEL'].iloc[0]
)
 
# Realocações - Suporte KCS (Contrato de fornecedor)
processado_erro.loc[
    (processado_erro['CAMPO'].str.contains('Contrato de fornecedor')),
    'RESPONSÁVEL'] = 'Suporte KCS'
 
# Realocações - Suporte KCS (TMS - ON LINE)
processado_erro.loc[
    (processado_erro['OBTS'] == 'ARGO(TMS)') &
    (processado_erro['Status Requisicao'] == 'ON LINE') &
    (processado_erro['Cliente'].str.contains('Argo It')),
    'RESPONSÁVEL'] = 'Suporte KCS'
 
# Realocações - Operações - CORP (MANUAL)
processado_erro.loc[
    (processado_erro['OBTS'] == 'MANUAL') &
    (processado_erro['RESPONSÁVEL'] == 'Suporte KCS'),
    'RESPONSÁVEL'] = 'Operações - CORP'
 
# Realocações - Suporte KCS (Latam, Gol - Bilhete incompleto)
processado_erro.loc[
    (processado_erro['OBTS'] == 'SABRE') &
    (processado_erro['CAMPO'].str.contains('Bilhete incompleto')) &
    (processado_erro['Fornecedor'].str.contains('Latam|Gol')),
    'RESPONSÁVEL'] = 'Suporte KCS'
   
# Realocações - Suporte KCS (GOVER)
processado_erro.loc[
    (processado_erro['OBTS'] == 'GOVER') &
    (processado_erro['Status Requisicao'] == 'ON LINE'),
    'RESPONSÁVEL'] = 'Suporte KCS'
 
# Realocações - Suporte KCS (Falta de Fornecedor - ONLINE - Carro)
processado_erro.loc[
    (processado_erro['Status Requisicao'] == 'ON LINE') &
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')) &
    (processado_erro['Serviço'] == 'Carro'),
    'RESPONSÁVEL'] = 'Suporte KCS'

# Realocações - Suporte KCS (Falta de Fornecedor - ONLINE - Lemontech)
processado_erro.loc[
    (processado_erro['Status Requisicao'] == 'ON LINE') &
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')) &
    (processado_erro['OBTS'] == 'LEMONTECH'),
    'RESPONSÁVEL'] = 'Suporte KCS'
 
# Realocações - Suporte KCS (Falta de Fornecedor - Offline)
processado_erro.loc[
    (processado_erro['Status Requisicao'] == 'OFF LINE') &
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')),
    'RESPONSÁVEL'] = 'Suporte KCS'
 
# Realocações - Suporte KCS (Falta de Fornecedor - ONLINE - Lemontech - Lojas Renner)
processado_erro.loc[
    (processado_erro['Status Requisicao'] == 'ON LINE') &
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')) &
    (processado_erro['OBTS'] == 'LEMONTECH') &
    (processado_erro['Grupo Empresarial'].str.contains('Grupo Lojas Renner')),
    'RESPONSÁVEL'] = 'Suporte KCS'
 
# Realocações - Suporte KCS (WS, CT, WT - SABRE)
processado_erro.loc[
    (processado_erro['OBTS'] == 'SABRE') &
    (processado_erro['Agente Emissão'].str.contains('WS') | processado_erro['Agente Emissão'].str.contains('CT') | processado_erro['Agente Emissão'].str.contains('WT')),
    'RESPONSÁVEL'] = 'Suporte KCS'
 
# Realocação - Suporte Benner (Falta de Fornecedor - Sabre - Hotel - ONLINE)
processado_erro.loc[
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')) &
    ((processado_erro['Agente Emissão'].str.contains('WS')) | (processado_erro['Agente Emissão'].str.contains('CT'))) &
    ((processado_erro['OBTS'] == 'SABRE') &
    (processado_erro['Serviço'] == 'Hotel')),['RESPONSÁVEL', 'CATEGORIA DE ERRO']
] = ['Suporte Benner', 'Sistêmico']

# Realocação - Operações - CORP (Falta de Fornecedor - KPMG - OFFLINE - Carro)
processado_erro.loc[
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')) &
    (processado_erro['Grupo Empresarial'].str.contains('Grupo Kpmg')) &
    (processado_erro['Serviço'] == 'Carro') &
    (processado_erro['Status Requisicao'] == 'OFF LINE') &
    (pd.to_datetime(processado_erro['Data Emissão'], dayfirst=True, errors='coerce') > pd.to_datetime('06/08/2024', dayfirst=True)),
    ['RESPONSÁVEL', 'CATEGORIA DE ERRO']
    ] = ['Operações - CORP', 'Qualidade dos dados']
 
# Realocação - Suporte KCS(Gover - Offline - Agente Emissão - Falta de informação Gerencial)
processado_erro.loc[
    (processado_erro['OBTS'] == 'GOVER') &
    (processado_erro['Status Requisicao'] == 'OFF LINE') &
    (processado_erro['Agente Emissão'].str.contains('Gover')) &
    (processado_erro['CAMPO'].str.contains('Falta de informação Gerencial')),
    'RESPONSÁVEL'] = 'Suporte KCS'
 
# Realocação - Suporte KCS(OBTS - Falta de informação Gerencial)
processado_erro.loc[
    (processado_erro['OBTS'].str.contains('ARGO|TMS|GOVER|LEMONTECH')) &
    (processado_erro['CAMPO'].str.contains('Falta de informação Gerencial')),
    'RESPONSÁVEL'] = 'Suporte KCS'
 
# Realocação - Suporte Benner (Falta de Fornecedor - GOVER - Hotel)
processado_erro.loc[
    (processado_erro['OBTS'] == 'GOVER') &
    (processado_erro['Serviço'] == 'Hotel') &
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')),
    ['RESPONSÁVEL', 'CATEGORIA DE ERRO']
] = ['Suporte Benner', 'Sistêmico']

# Realocação - Suporte Benner (Falta de Fornecedor - GOVER - Carro - ONLINE)
processado_erro.loc[
    (processado_erro['OBTS'] == 'GOVER') &
    (processado_erro['Serviço'] == 'Carro') &
    (processado_erro['Status Requisicao'] == 'ON LINE') &
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')),['RESPONSÁVEL', 'CATEGORIA DE ERRO']
] = ['Suporte Benner', 'Sistêmico']
 
# Realocação - Suporte Benner (Falta de Fornecedor - Argo - Hotel - ONLINE)
processado_erro.loc[
    (processado_erro['OBTS'] == 'ARGO(TMS)') &
    (processado_erro['Serviço'] == 'Hotel') &
    (processado_erro['Status Requisicao'] == 'ON LINE') &
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')),['RESPONSÁVEL', 'CATEGORIA DE ERRO']
] = ['Suporte Benner', 'Sistêmico']

# Realocação - Suporte Benner (Falta de Fornecedor - GOVER - Hotel)
processado_erro.loc[
    (processado_erro['OBTS'] == 'GOVER') &
    (processado_erro['Serviço'] == 'Hotel') &
    (processado_erro['CAMPO'].str.contains('Falta de Fornecedor')),['RESPONSÁVEL', 'CATEGORIA DE ERRO']
] = ['Suporte Benner', 'Sistêmico']

# Realocação - Agencia Mercurio York
processado_erro.loc[
    (processado_erro['Cliente'] == 'Agencia Mercurio York'),
    ['OBTS', 'RESPONSÁVEL']] = ['MANUAL', 'Operações Mercurio York']
 
# Realocação - KONTRIP
processado_erro.loc[
    (processado_erro['OBTS'].str.contains('KONTRIP')),
    'RESPONSÁVEL'] = 'Operações - KONTRIP'

# Realocação - ZUPPER
processado_erro.loc[
    (processado_erro['OBTS'].str.contains('ZUPPER')),
    'RESPONSÁVEL'] = 'Operações - ZUPPER'
 
# Realocação - Suporte Benner (Cliente FEE no POS)
processado_erro.loc[
    (processado_erro['Cliente Fee POS'] == 'Cliente FEE no POS') &
    (processado_erro['CAMPO'].str.contains('Pagamento não permitido para cobrança')) &
    (processado_erro['Forma Pagamento'] == 'Pagamento direto'),
    ['RESPONSÁVEL', 'CATEGORIA DE ERRO']
] = ['Suporte Benner', 'Sistêmico']
 
# Realocação - Usando planilha de realocação - por localizador
realocacao_handle = processado_erro.merge(realocacao[['Handle ACC', 'Equipe ATUAL']], on='Handle ACC', how='left')
processado_erro.loc[realocacao_handle['Equipe ATUAL'].notnull(), 'RESPONSÁVEL'] = realocacao_handle['Equipe ATUAL']
 
# Realocação- Conciliação aérea (Duplicidade de RLOC ou Bilhete duplicado)
processado_erro.loc[
    (processado_erro['Serviço'].str.contains('Aéreo')) &
    ((processado_erro['CAMPO'].str.contains('Duplicidade de RLOC')) |
     (processado_erro['CAMPO'].str.contains('Bilhete duplicado'))),
    'RESPONSÁVEL'] = 'Conciliação aérea'

# Realocação - Central de Emissão (Markup) por Agente Emissão
agentes = [
    'Gessiane Santos Silva', 'Clayton Alves de Rezende', 'Icaro Gabriel Pimentel Gomes Xavier', 
    'Carlos Alberto Theodoro da Silva', 'Mylena Mendonca Santos da Silva', 'Wellington Ribeiro da Silva', 
    'Itamar de Souza', 'Flavia Constanzi do Nascimento', 'Ana Paula Costa Feitosa', 'Elisete Ferraz de Almeida','Beatrys Ferreira Rocha'
]

processado_erro.loc[
    processado_erro['Agente Emissão'].isin(agentes) | processado_erro['Agente Criação'].isin(agentes),
    ['RESPONSÁVEL', 'CATEGORIA DE ERRO']
] = ['Central de Emissão', 'Qualidade dos dados']


# Realocações - Suporte KCS (Bilhete UNDEFINED)
processado_erro.loc[
    (processado_erro['OBTS'] == 'SABRE') &
    (processado_erro['Mensagem Erro'].str.contains('UNDEFINED')),
    'RESPONSÁVEL'] = 'Suporte KCS'

# Realocações - Suporte KCS (The INSERT statement conflicted with the FOREIGN KEY)
processado_erro.loc[
    (processado_erro['Mensagem Erro'].str.contains(
        'The INSERT statement conflicted with the FOREIGN KEY')),
    ['CAMPO', 'ORIGEM DO ERRO', 'TIPO DE ERRO', 'CATEGORIA DE ERRO', 'RESPONSÁVEL']
] = ['Falta de informação Gerencial', 'Mais de um campo não preenchido', 'Dados do Fornecedor', 'Sistêmico', 'Suporte KCS']
 
# Realocação - Conciliação aérea (Esta accounting está conciliada no BSP. Bilhete)
processado_erro.loc[
    (processado_erro['Mensagem Erro'].str.contains(
        'Esta accounting está conciliada no BSP. Bilhete:')),
    'RESPONSÁVEL'] = 'Conciliação aérea'
 
# Realocação - Central de Emissão (Markup)
processado_erro['Markup'] = processado_erro['Markup'].astype(float)
processado_erro.loc[processado_erro['Markup']
                    > 0, 'RESPONSÁVEL'] = 'Central de Emissão'

# Realocação - INOVENTS (Codigo Evento)
processado_erro.loc[
    processado_erro['Canal de Vendas'].str.contains('INOVENTS'),
    'RESPONSÁVEL'] = 'INOVENTS'

# Realocações - EMPRESA - KONTIK BUSINESS TRAVEL
processado_erro.loc[
    processado_erro['RESPONSÁVEL'].str.contains('Operações - CORP') | processado_erro['RESPONSÁVEL'].str.contains('Operações - FCM') |
    processado_erro['RESPONSÁVEL'].str.contains('Suporte KCS'),
    'EMPRESA'] = 'KONTIK BUSINESS TRAVEL'

# Realocações - EMPRESA - ZUPPER VIAGENS
processado_erro.loc[
    processado_erro['OBTS'].str.contains('ZUPPER'),
    'EMPRESA'] = 'ZUPPER VIAGENS'

# Realocações - EMPRESA - INOVENTS
processado_erro.loc[
    processado_erro['RESPONSÁVEL'].str.contains('INOVENTS'),
    'EMPRESA'] = 'INOVENTS'

# Realocações - EMPRESA - GRUPO KONTIK
processado_erro.loc[
    processado_erro['RESPONSÁVEL'].str.contains('Central de Emissão') | processado_erro['RESPONSÁVEL'].str.contains('Conciliação aérea') |
    processado_erro['RESPONSÁVEL'].str.contains('Suporte Benner'),
    'EMPRESA'] = 'GRUPO KONTIK'

# Realocações - EMPRESA - KONTRIP VIAGENS
processado_erro.loc[
    processado_erro['OBTS'].str.contains('KONTRIP'),
    'EMPRESA'] = 'KONTRIP VIAGENS'

# Realocações - EMPRESA - K-CLUB
processado_erro.loc[
    processado_erro['RESPONSÁVEL'].str.contains('K-CLUB'),
    'EMPRESA'] = 'KTK'

# Realocação - Suporte Benner (Não foi possível definir o Local de destino! - Suporte Benner)
processado_erro.loc[
    (processado_erro['Mensagem Erro'].str.contains('Não foi possível definir o Local de destino!')),
     'RESPONSÁVEL'] = 'Suporte Benner'

# Verificação de duplicidade(Gover)
cont_localizador = processado_erro['Localizadora'].tolist()
 
for row in range(len(processado_erro)):
    if cont_localizador.count(processado_erro['Localizadora'][row]) > 10 and \
        'GOVER' in processado_erro['OBTS'][row]:
        processado_erro.at[row, 'CAMPO'] = 'Falha no processo de integração'
        processado_erro.at[row,
                           'ORIGEM DO ERRO'] = 'Vendas duplicadas'
        processado_erro.at[row, 'TIPO DE ERRO'] = 'Sistema'
        processado_erro.at[row, 'RESPONSÁVEL'] = 'Suporte Benner'
        processado_erro.at[row, 'CATEGORIA DE ERRO'] = 'Sistêmico'
    
        if 'Vendas duplicadas' in str(processado_erro.at[row, 'ORIGEM DO ERRO']):
            print(f'\033[1;31m-Verifique o localizador "{processado_erro.at[row, "Localizadora"]}"'
                  f' e requisição "{processado_erro.at[row, "Requisição"]}",'
                  f' feitas pelo consultor "{processado_erro.at[row, "Agente Emissão"]}"\033[m')
            break

# Salvar o arquivo original
processado_erro.to_excel(data_path + 'Processado Erro.xlsx', index=False)
 
# Seleção de colunas
colunas = {
    'A': 'Handle PNR', 'B': 'Handle ACC', 'C': 'Sequencia', 'D': 'Data Inclusão',
    'E': 'Data Alteração', 'F': 'Aging Alteração', 'G': 'Aging Inclusão', 'H': 'Localizadora', 'I': 'OBT', 'J': 'Pax',
    'K': 'Agente Emissão', 'L': 'Agente Criação', 'M': 'Data Emissão', 'N': 'Requisição', 'O': 'Local Retirada',
    'P': 'Status Requisicao', 'Q': 'Forma Pagamento', 'R': 'Forma Recebimento',
    'S': 'Serviço', 'T': 'Cancelado', 'U': 'Grupo Empresarial', 'V': 'Cliente',
    'W': 'Cliente Fee POS', 'X': 'Fornecedor', 'Y': 'Bilhete','Z': 'Canal de Vendas',
    'AA': 'Codigo Evento','AB': 'Tarifa', 'AC': 'Taxa', 'AD': 'Outras Taxas',
    'AE': 'Taxa DU', 'AF': 'Taxa BR', 'AG': 'Taxa Extra', 'AH': 'Fee',
    'AI': 'Observação', 'AJ': 'Mensagem Erro', 'AK': 'OBTS', 'AL': 'CAMPO',
    'AM': 'ORIGEM DO ERRO', 'AN': 'TIPO DE ERRO', 'AO': 'CATEGORIA DE ERRO',
    'AP': 'EMPRESA','AQ': 'RESPONSÁVEL'
}
 
# Inserção de colunas
for col, nome in colunas.items():
    relatorio_base[col + '1'] = nome
 
# Inserção de valores
for row in range(len(processado_erro)):
    for col, nome in colunas.items():
        relatorio_base[col + str(row + 2)] = processado_erro[nome][row]
 
# Personalização de células
def personalizacao(relatorio):
    # Definindo cores para o cabeçalho
    colunas_padrao = PatternFill(start_color="591F6A", end_color="591F6A", fill_type="solid")  # ROXO
    colunas_adicionais = PatternFill(start_color="18F194", end_color="18F194", fill_type="solid")  # VERDE
    cor_da_fonte_1 = Font(color="18F194", bold=True) # cor forte VERDE
    cor_da_fonte_2 = Font(color="591F6A", bold=True) # cor fonte ROXO
    
    # Alinhamento à esquerda
    alinhamento_esquerda = Alignment(horizontal="left")

    # Aplicar a formatação aos cabeçalhos
    for coluna in range(1, 44):  # Colunas de A até AS
        letra_coluna = relatorio.cell(row=1, column=coluna).column_letter
        cabecalho = relatorio[letra_coluna + '1']
        cabecalho.fill = colunas_padrao
        cabecalho.font = cor_da_fonte_1
        # cabecalho.value = cabecalho.value.upper()  # Convertendo para maiúsculo
        
        if letra_coluna in ['F','G','AK','AL','AM', 'AN','AO','AP','AQ']:
            cabecalho.fill = colunas_adicionais
            cabecalho.font = cor_da_fonte_2
    
    # Alinhamento à esquerda nas células do corpo do relatório
    for row in relatorio.iter_rows(min_row=2, min_col=1, max_col=43):  # Ajuste para percorrer todas as colunas
        for cell in row:
            cell.alignment = alinhamento_esquerda  # Alinhando células à esquerda

    # Definindo cores para aging de dias
    for row in relatorio.iter_rows(min_row=2, min_col=6, max_col=8):
        cell_value = row[0].value
        if cell_value in ['16 a 23 dias', '24 a 31 dias', '31 dias ou +']:
            row[0].font = Font(color="FF0000")
            row[1].font = Font(color="FF0000")

personalizacao(relatorio_base)
# Salvar relatorio base
wb.save(data_path + 'Relatorio - Dash.xlsx')

# Carregar a planilha original
relatorio_dash = pd.read_excel(data_path + 'Relatorio - Dash.xlsx', sheet_name='Processado Erro - BASE')
data_path = r'PROCESSADO ERRO\Analise de Dados\\'

empresas = relatorio_dash['EMPRESA'].unique()

# Defina o caminho do diretório onde os arquivos serão salvos
output_folder = data_path + 'EMPRESAS'

# Verifique se o diretório existe; se não, crie-o
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

for empresa in empresas:
    # Criar uma nova planilha para cada empresa
    df_empresa = relatorio_dash[relatorio_dash['EMPRESA'] == empresa]
    output_file = os.path.join(output_folder, f'Relatorio - {empresa}.xlsx')

    # Salvando o DataFrame da empresa em um arquivo Excel
    df_empresa.to_excel(output_file, index=False)

    # Abrindo o arquivo Excel para aplicar formatação
    wb = load_workbook(output_file)
    ws = wb.active  # Pegando a primeira aba
    
    # Aplicando a personalização
    personalizacao(ws)
    
    # Salvando as mudanças no arquivo Excel
    wb.save(output_file)

    print(f'Relatorio - {empresa} - Salvo com sucesso!')
print()
print('\033[1;32m-Relatório gerado com sucesso!\033[m')
